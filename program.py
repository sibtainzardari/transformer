import math
import openpyxl

print("Enter the Primary Voltage Vp in volt (V)")
Vp = float(input())

print("Enter the Secondary Voltage Vs in volt (V)")
Vs = float(input())

print("Enter the Secondary Current Is in ampere (A)")
Is = float(input())

print("Enter the value of K through the table")
K = float(input())

print("Enter the value of frequency f in Hz")
f = float(input())

print("Enter the value of Magnetic flux density Bmax in wb/m²")
Bmax = float(input())

print("Enter the value of current density J in A/m²")
J = float(input())

print("Enter the value of paper width Wp in mm")
Wp = float(input())

Ep = Vp
Es = Vs
Ip = (Vs * Is) / Vp
Ss = ((Vs * Is) / 1000)
print("Power of secondary side (Ss) is: ", Ss, "KVA")
Sp = Ss

Q = math.sqrt(Ss)
Et = round(K * Q, 3)
print("Emf per turn of the coil Et is: ", Et, "V/T") 
Np = int(Vp / Et)
print("Number of primary turns per phase Np is: ", Np, "turns")
Ns = int((Vs * Np) / Vp)
print("Number of secondary turns per phase Ns is: ", Ns, "turns")
Ab = int(Ep / (4.44 * f * Bmax * Np * 0.98 * pow(10, -6)))
print("Calculating the area of the bobbin Ab is: ", Ab, "mm²")
Wb1 = int(math.sqrt(Ab))
print("One side length of the bobbin Wb1 is: ", Wb1, "mm")

Apc = round(Ip / J, 3)
print("The cross-section of primary conductor is: ", Apc, "mm²")
Asc = round(Is / J, 3)
print("The cross-section of secondary conductor is: ", Asc, "mm²")

# Load the Excel workbook for Sankey Number
sankey_workbook = openpyxl.load_workbook('sankey.xlsx')
sankey_worksheet = sankey_workbook['Sheet1']

# Get the input value for Wl1
input_value1 = Wb1

closest_rows = []
closest_abs_diff = float('inf')

for row in sankey_worksheet.iter_rows(min_row=2, max_row=sankey_worksheet.max_row, min_col=2, max_col=2):
    cell_value = row[0].value
    if cell_value is not None and isinstance(cell_value, (int, float)):
        abs_diff = abs(cell_value - input_value1)
        if abs_diff <= closest_abs_diff:
            if abs_diff < closest_abs_diff:
                closest_rows.clear()
            closest_abs_diff = abs_diff
            closest_rows.append([cell.value for cell in sankey_worksheet[row[0].row]])

if closest_rows:
    print("Rows of the closest former area Af:")
    for row_values in closest_rows:
        print(*row_values, sep=', ')

WH = float(input("Enter the value of window height WH in mm: "))

closest_row = None
closest_abs_diff = float('inf')

for row_values in closest_rows:
    cell_value = row_values[2]
    abs_diff = abs(cell_value - WH)
    if abs_diff < closest_abs_diff:
        closest_abs_diff = abs_diff
        closest_row = row_values

if closest_row is not None:
    print("The closest row for Af and WH:")
    print(*closest_row, sep=', ')

Wb1 = closest_row[1]
WW = closest_row[3]
Sankey_num = closest_row[4]
Df = closest_row[5]

print(f'The available one side of the bobbin length Wb1 is: {Wb1}', "mm")
Wb2 = int(Ab/Wb1)
print(f'The available second side of the bobbin length Wb2 is: {Wb2}', "mm")

print(f'The Sankey Number is: {Sankey_num}')
print(f'The window width of the former WW is: {WW}', "mm")

# Load the Excel workbook for Conductor
conductor_workbook = openpyxl.load_workbook('conductor.xlsx')
conductor_worksheet = conductor_workbook['Sheet1']

# Get the input value for Apc
input_value1 = Apc

closest_rows = []
closest_abs_diff = float('inf')

for row in conductor_worksheet.iter_rows(min_row=2, max_row=conductor_worksheet.max_row, min_col=2, max_col=2):
    cell_value = row[0].value
    if cell_value is not None and isinstance(cell_value, (int, float)):
        abs_diff = abs(cell_value - input_value1)
        if abs_diff <= closest_abs_diff:
            if abs_diff < closest_abs_diff:
                closest_rows.clear()
            closest_abs_diff = abs_diff
            closest_rows.append([cell.value for cell in conductor_worksheet[row[0].row]])

if closest_rows:
    print("Rows of the closest area of primary conductor:")
    for row_values in closest_rows:
        print(*row_values, sep=', ')

Apc = row_values[1]
SWG = row_values[3]
Dpc = row_values[2]

print(f'The area of primary conductor from the table Apc is: {Apc}', "mm²")
print(f'The diameter of primary conductor Dpc is: {Dpc}', "mm")
print(f'The wire guage SWG1 of primary conductor is: {SWG}')

# Get the input value for Asc
input_value2 = Asc

closest_rows = []
closest_abs_diff = float('inf')

for row in conductor_worksheet.iter_rows(min_row=2, max_row=conductor_worksheet.max_row, min_col=2, max_col=2):
    cell_value = row[0].value
    if cell_value is not None and isinstance(cell_value, (int, float)):
        abs_diff = abs(cell_value - input_value2)
        if abs_diff <= closest_abs_diff:
            if abs_diff < closest_abs_diff:
                closest_rows.clear()
            closest_abs_diff = abs_diff
            closest_rows.append([cell.value for cell in conductor_worksheet[row[0].row]])

if closest_rows:
    print("Row of the closest area of secondary conductor:")
    for row_values in closest_rows:
        print(*row_values, sep=', ')

Asc = row_values[1]
SWG = row_values[3]
Dsc = row_values[2]

print(f'The area of secondary conductor from the table Asc is: {Asc}', "mm²")
print(f'The diameter of secondary conductor Dsc is: {Dsc}', "mm")
print(f'The wire guage SWG2 of secondary conductor is: {SWG}')

Aapca = (3.14*Dpc*Dpc)/4
NOTpw = int(WH/Dpc)
print(f'The number of turns of primary winding NOTpw is: {NOTpw}', "T")
NLp = round(Np/NOTpw, 1)
NLpw = math.ceil(NLp)
print(f'The number of layers in primary winding NLpw is: {NLpw}')

Wb21 = (2*NLpw*Dpc)+Wb1+2*Wp*(NLpw-1)+2
Wb22 = (2*NLpw*Dpc)+Wb2+2*Wp*(NLpw-1)+2
Perisw = 2*(Wb21+Wb22)
Aasca = (3.14*Dsc*Dsc)/4
NOTsw = int(WH/Dsc)
print(f'The number of turns of secondary winding NOTsw is: {NOTsw}', "T")
NLs = round(Ns/NOTsw, 1)
NLsw = math.ceil(NLs)
print(f'The number of layers in secondary winding NLsw is: {NLsw}')

IDp = Wb1
ODp = IDp+2*Dpc*NLpw+2*NLpw*Wp
ADp = (IDp+ODp)/2
Bp = (ODp-IDp)/2

IDs = ODp+1
ODs = IDs+2*NLsw*Dsc+2*NLsw*Wp
Bs = (ODs-IDs)/2
ADs = (IDs+ODs)/2

# Print Window gap Wg:
Cd = Dpc * NLpw + Dsc * NLsw + Wp * (NLpw + NLsw - 2) + 5
Wg = WW - Cd
Wg = round(Wg, 1)
print("The value of Wg is:", Wg)

# Check condition and adjust Dpc and Dsc
if 0 < Wg < 2:
    print("The value of Wg is:", Wg)
elif Wg < 0:
    while Wg < 0:
        Dpc -= 0.5
        Dsc -= 0.5
        Cd = Dpc * NLpw + Dsc * NLsw + Wp * (NLpw + NLsw - 2) + 5
        Wg = WW - Cd
        print("The new value of Wg is:", Wg)
else:
    while Wg > 2:
        Dpc += 0.5
        Dsc += 0.5
        Cd = Dpc * NLpw + Dsc * NLsw + Wp * (NLpw + NLsw - 2) + 5
        Wg = WW - Cd
        print("The new value of Wg is:", Wg)

# Print the adjusted values
print("Adjusted Dpc:", Dpc)
print("Adjusted Dsc:", Dsc)
        
Peripw = 2*(Wb1+Wb2)
Lp = ((NOTpw * NLpw * Peripw) + (4 * NOTpw * NLpw * Dpc * (NLpw + 1)) + (4 * NLpw * Wp * (NLpw - 1)))
    
Lpw = int(Lp/1000)
print(f'value of Lpw is: {Lpw}', "m")

p1 = 0.0177 #Value of Resistivity in meter

Rpw = round((p1*Lpw)/Aapca, 3)
print(f'The primary winding resistance Rpw is: {Rpw}', "ohm")

Ls = ((NOTsw * NLsw * Perisw) + (4 * NOTsw * NLsw * Dsc * (NLsw + 1)) + (4 * NLsw * Wp * (NLsw - 1)))
Lsw = int(Ls/1000)
print(f'value of Lsw is: {Lsw}', "m")
Rsw = round((p1*Lsw)/Aasca, 3)
print(f'The secondary winding resistance Rsw is: {Rsw}', "ohm")

Rpw1 = Rpw*(275+75)/(273+25)

Rsw1 = Rsw*(275+75)/(273+25)

#Impedence Calculation
p = 2*3.14*f*4*3.14*pow(10, -7)*Np*Np
q = ADs/WH
r = 1+(Bp+Bs)/3
Xp = (p+q+r)/1000
Zb = (Vp*Vp)/(Sp*1000)
PUlr = Xp/Zb
Tr = Np/Ns
RRtp = Rpw+(Rsw*Tr*Tr)
PUr = RRtp/Zb
Zpu = math.sqrt(PUlr*PUlr+PUr*PUr)
Z = Zpu*100
print("The value of %Impedence is:",Z,"Ohm")
